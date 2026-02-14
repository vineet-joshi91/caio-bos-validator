# -*- coding: utf-8 -*-
"""
Created on Sat Feb 14 16:39:49 2026

@author: Vineet
"""

# api/services/document_service.py
"""
Document processing service
Handles text extraction from various file formats
"""
import logging
from typing import Dict, Tuple, Any

logger = logging.getLogger(__name__)

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    import docx
except Exception:
    docx = None

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import pytesseract
    from PIL import Image
except Exception:
    pytesseract = None
    Image = None


def extract_text_from_upload(filename: str, data: bytes) -> str:
    """
    Extract text from uploaded file.
    Supports: PDF, DOCX, XLSX, TXT, CSV
    
    Args:
        filename: Original filename
        data: File bytes
    
    Returns:
        Extracted text as string
    """
    lower_name = filename.lower()
    
    # PDF
    if lower_name.endswith(".pdf"):
        return _extract_from_pdf(data)
    
    # Word document
    elif lower_name.endswith(".docx"):
        return _extract_from_docx(data)
    
    # Excel
    elif lower_name.endswith((".xlsx", ".xls")):
        return _extract_from_excel(data)
    
    # Plain text
    elif lower_name.endswith(".txt"):
        return data.decode("utf-8", errors="ignore")
    
    # CSV
    elif lower_name.endswith(".csv"):
        return data.decode("utf-8", errors="ignore")
    
    else:
        logger.warning(f"Unsupported file type: {filename}")
        return ""


def extract_text_with_meta(filename: str, data: bytes) -> Tuple[str, Dict[str, Any]]:
    """
    Extract text with metadata about extraction quality.
    
    Returns:
        (extracted_text, metadata_dict)
    """
    text = extract_text_from_upload(filename, data)
    
    meta = {
        "filename": filename,
        "size_bytes": len(data),
        "text_length": len(text),
        "chosen_method": _get_extraction_method(filename),
        "quality_flags": _get_quality_flags(filename, text, data),
    }
    
    return text, meta


def _extract_from_pdf(data: bytes) -> str:
    """Extract text from PDF bytes"""
    import io
    
    if not PdfReader:
        return ""
    
    try:
        pdf = PdfReader(io.BytesIO(data))
        text_parts = []
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
        return ""


def _extract_from_docx(data: bytes) -> str:
    """Extract text from DOCX bytes"""
    import io
    
    if not docx:
        return ""
    
    try:
        doc = docx.Document(io.BytesIO(data))
        text_parts = [para.text for para in doc.paragraphs]
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"DOCX extraction failed: {e}")
        return ""


def _extract_from_excel(data: bytes) -> str:
    """Extract text from Excel bytes"""
    import io
    
    if not openpyxl:
        return ""
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return ""


def _get_extraction_method(filename: str) -> str:
    """Determine which extraction method was used"""
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return "pypdf"
    elif lower.endswith(".docx"):
        return "python-docx"
    elif lower.endswith((".xlsx", ".xls")):
        return "openpyxl"
    elif lower.endswith((".txt", ".csv")):
        return "direct-decode"
    return "unknown"


def _get_quality_flags(filename: str, text: str, data: bytes) -> list:
    """Generate quality warning flags for extraction"""
    flags = []
    
    # Check if very little text was extracted from PDF
    if filename.lower().endswith(".pdf") and len(text) < 200 and len(data) > 10000:
        flags.append("LOW_TEXT_PDF")
    
    # Check for likely pricing/quotation docs (common issue)
    lower_text = text.lower()
    if any(word in lower_text for word in ["quote", "quotation", "pricing", "invoice"]):
        if len(text) < 500:  # Very short for a pricing doc
            flags.append("LIKELY_QUOTE_PRICING_NOT_EXTRACTED")
    
    return flags